import asyncio
import logging
from io import BytesIO
from telegram import Update
from telegram.ext import ContextTypes
from telegram.constants import ParseMode

from services.cloudconvert import CloudConvertService
from services.file_handler import FileHandler
from services.database import Database
from bot import messages
from bot.keyboards import *
from config.settings import MAX_FILE_SIZE, ERROR_MESSAGES, CLAUDE_ENABLED

# Импортируем Claude сервис только если он включен
if CLAUDE_ENABLED:
    from services.claude_service import ClaudeService

logger = logging.getLogger(__name__)

class BotHandlers:
    def __init__(self):
        self.cloudconvert = CloudConvertService()
        self.file_handler = FileHandler()
        self.db = Database()
        
        # Инициализируем Claude только если он включен
        if CLAUDE_ENABLED:
            self.claude = ClaudeService()
        else:
            self.claude = None
    
    async def start_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик команды /start"""
        user = update.effective_user
        
        # Логируем старт
        self.db.log_operation(user.id, user.username, "start", "completed")
        
        await update.message.reply_text(
            messages.START_MESSAGE,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=get_start_keyboard()
        )
    
    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик команды /help"""
        await update.message.reply_text(
            messages.HELP_MESSAGE,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=get_help_keyboard()
        )
    
    async def convert_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик команды /convert"""
        await update.message.reply_text(
            messages.CONVERT_MESSAGE,
            parse_mode=ParseMode.MARKDOWN
        )
    
    async def status_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик команды /status"""
        user_id = update.effective_user.id
        active_task = self.db.get_active_task(user_id)
        
        if not active_task:
            await update.message.reply_text(
                messages.STATUS_NO_ACTIVE_TASKS,
                parse_mode=ParseMode.MARKDOWN
            )
        else:
            status_text = messages.STATUS_ACTIVE_TASK.format(
                filename=active_task['file_name'],
                status="🔄 Обработка",
                started_at=active_task['created_at']
            )
            await update.message.reply_text(
                status_text,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=get_status_keyboard()
            )
    
    async def handle_document(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик получения документа"""
        user = update.effective_user
        document = update.message.document
        
        # Проверяем лимит запросов
        if not self.db.check_user_rate_limit(user.id):
            await update.message.reply_text(
                messages.ERROR_RATE_LIMIT,
                parse_mode=ParseMode.MARKDOWN
            )
            return
        
        # Проверяем размер файла
        if document.file_size > MAX_FILE_SIZE:
            size_mb = round(document.file_size / (1024*1024), 2)
            await update.message.reply_text(
                messages.ERROR_FILE_TOO_LARGE.format(size=size_mb),
                parse_mode=ParseMode.MARKDOWN
            )
            return
        
        # Проверяем формат файла
        if not document.file_name.lower().endswith('.pdf'):
            await update.message.reply_text(
                messages.ERROR_INVALID_FORMAT,
                parse_mode=ParseMode.MARKDOWN
            )
            return
        
        # Начинаем обработку файла
        await self._process_file(update, context, document)
    
    async def _process_file(self, update: Update, context: ContextTypes.DEFAULT_TYPE, document):
        """Обработка PDF файла"""
        user = update.effective_user
        
        # Отправляем сообщение о начале обработки
        size_mb = round(document.file_size / (1024*1024), 2)
        processing_msg = await update.message.reply_text(
            messages.PROCESSING_START.format(filename=document.file_name, size=size_mb),
            parse_mode=ParseMode.MARKDOWN
        )
        
        try:
            # Логируем операцию
            operation_id = self.db.log_operation(
                user.id, user.username, "conversion", "processing",
                document.file_name, document.file_size
            )
            
            # Скачиваем файл
            file = await context.bot.get_file(document.file_id)
            file_data = await file.download_as_bytearray()
            
            # Сохраняем в базу активную задачу
            self.db.save_active_task(user.id, f"temp_{user.id}", document.file_name)
            
            # Обновляем сообщение
            await processing_msg.edit_text(
                messages.CONVERSION_MESSAGES['uploading'],
                parse_mode=ParseMode.MARKDOWN
            )
            
            # Конвертируем файл
            await processing_msg.edit_text(
                messages.CONVERSION_MESSAGES['processing'],
                parse_mode=ParseMode.MARKDOWN
            )
            
            # Выполняем конвертацию
            converted_data = await self.cloudconvert.convert_pdf_to_xlsx(
                bytes(file_data), document.file_name
            )
            
            if converted_data:
                # Успешная конвертация
                await processing_msg.edit_text(
                    messages.CONVERSION_SUCCESS,
                    parse_mode=ParseMode.MARKDOWN
                )
                
                # Анализ и улучшение качества с помощью Claude AI (если включен)
                enhanced_data = converted_data
                enhancement_stats = None
                
                try:
                    from config.settings import CLAUDE_ENABLED
                    from services.text_enhancer import TextEnhancer
                    
                    if CLAUDE_ENABLED:
                        await processing_msg.edit_text(
                            messages.CONVERSION_MESSAGES['enhancing'],
                            parse_mode=ParseMode.MARKDOWN
                        )
                        
                        enhancer = TextEnhancer()
                        
                        # Анализируем оригинальный файл
                        try:
                            import openpyxl
                            from io import BytesIO
                            
                            workbook = openpyxl.load_workbook(BytesIO(converted_data))
                            original_text = ""
                            for sheet in workbook.worksheets:
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value and isinstance(cell.value, str):
                                            original_text += cell.value + " "
                            
                            # Улучшаем файл
                            enhanced_data = await enhancer.process_xlsx_file(converted_data, document.file_name)
                            
                            # Анализируем улучшенный файл
                            if enhanced_data != converted_data:
                                workbook_enhanced = openpyxl.load_workbook(BytesIO(enhanced_data))
                                enhanced_text = ""
                                for sheet in workbook_enhanced.worksheets:
                                    for row in sheet.iter_rows():
                                        for cell in row:
                                            if cell.value and isinstance(cell.value, str):
                                                enhanced_text += cell.value + " "
                                
                                enhancement_stats = enhancer.get_enhancement_stats(original_text, enhanced_text)
                                logger.info(f"Text enhancement stats: {enhancement_stats}")
                            
                        except Exception as analysis_error:
                            logger.error(f"Error in text analysis: {analysis_error}")
                            
                except ImportError:
                    logger.info("Claude AI enhancement not available")
                except Exception as claude_error:
                    logger.error(f"Error in Claude AI enhancement: {claude_error}")
                
                # Отправляем файл
                await processing_msg.edit_text(
                    messages.CONVERSION_MESSAGES['finalizing'],
                    parse_mode=ParseMode.MARKDOWN
                )
                
                # Генерируем имя XLSX файла
                xlsx_name = document.file_name.replace('.pdf', '.xlsx')
                
                # Формируем caption с информацией о качестве
                caption = "✅ Конвертация завершена успешно!"
                
                if enhancement_stats and enhancement_stats['improvement'] > 0:
                    caption += f"\n🤖 Текст улучшен с помощью Claude AI"
                    caption += f"\n📈 Качество: {enhancement_stats['original_score']:.0f}% → {enhancement_stats['enhanced_score']:.0f}%"
                    if enhancement_stats['ukrainian_chars_fixed'] > 0:
                        caption += f"\n🔧 Исправлено украинских символов: {enhancement_stats['ukrainian_chars_fixed']}"
                    if enhancement_stats['ocr_errors_fixed'] > 0:
                        caption += f"\n🔧 Исправлено OCR ошибок: {enhancement_stats['ocr_errors_fixed']}"
                elif CLAUDE_ENABLED:
                    caption += "\n✅ Качество проверено - улучшения не требуются"
                
                # Отправляем файл
                await update.message.reply_document(
                    document=BytesIO(enhanced_data),
                    filename=xlsx_name,
                    caption=caption,
                    reply_markup=get_success_keyboard()
                )
                
                # Обновляем статус в базе
                self.db.update_operation_status(operation_id, "completed")
                
            else:
                # Ошибка конвертации
                await processing_msg.edit_text(
                    messages.ERROR_CONVERSION_FAILED,
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=get_error_keyboard()
                )
                
                # Обновляем статус в базе
                self.db.update_operation_status(operation_id, "error", "Conversion failed")
        
        except Exception as e:
            logger.error(f"Error processing file: {e}")
            
            await processing_msg.edit_text(
                messages.ERROR_API_UNAVAILABLE,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=get_error_keyboard()
            )
            
            # Обновляем статус в базе
            if 'operation_id' in locals():
                self.db.update_operation_status(operation_id, "error", str(e))
        
        finally:
            # Удаляем активную задачу
            self.db.remove_active_task(user.id)
    
    async def handle_callback_query(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик callback кнопок"""
        query = update.callback_query
        await query.answer()
        
        data = query.data
        
        if data == "start_convert":
            await query.edit_message_text(
                messages.CONVERT_MESSAGE,
                parse_mode=ParseMode.MARKDOWN
            )
        
        elif data == "help":
            await query.edit_message_text(
                messages.HELP_MESSAGE,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=get_help_keyboard()
            )
        
        elif data == "status":
            user_id = query.from_user.id
            active_task = self.db.get_active_task(user_id)
            
            if not active_task:
                await query.edit_message_text(
                    messages.STATUS_NO_ACTIVE_TASKS,
                    parse_mode=ParseMode.MARKDOWN
                )
            else:
                status_text = messages.STATUS_ACTIVE_TASK.format(
                    filename=active_task['file_name'],
                    status="🔄 Обработка",
                    started_at=active_task['created_at']
                )
                await query.edit_message_text(
                    status_text,
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=get_status_keyboard()
                )
        
        elif data == "cancel_task":
            user_id = query.from_user.id
            self.db.remove_active_task(user_id)
            await query.edit_message_text(
                messages.CANCEL_SUCCESS,
                parse_mode=ParseMode.MARKDOWN
            )
        
        elif data == "retry_convert":
            await query.edit_message_text(
                messages.CONVERT_MESSAGE,
                parse_mode=ParseMode.MARKDOWN
            )
        
        elif data == "convert_another":
            await query.edit_message_text(
                messages.CONVERT_MESSAGE,
                parse_mode=ParseMode.MARKDOWN
            )
        
        elif data == "show_stats":
            stats = self.db.get_stats()
            stats_text = f"""
📊 **Статистика бота**

📈 **Всего операций:** {stats['total_operations']}
✅ **Успешных:** {stats['successful_operations']}
❌ **С ошибками:** {stats['error_operations']}
👥 **Уникальных пользователей:** {stats['unique_users']}
📊 **Успешность:** {stats['success_rate']:.1f}%
            """
            await query.edit_message_text(
                stats_text,
                parse_mode=ParseMode.MARKDOWN
            )
    
    async def handle_unknown_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик неизвестных команд"""
        await update.message.reply_text(
            messages.UNKNOWN_COMMAND,
            parse_mode=ParseMode.MARKDOWN
        ) 