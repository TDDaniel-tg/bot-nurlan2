import logging
import re
import asyncio
from typing import Optional, List, Dict, Tuple
from io import BytesIO
import openpyxl
import anthropic
from config.settings import CLAUDE_API_KEY, CLAUDE_MODEL, CLAUDE_ENABLED

logger = logging.getLogger(__name__)

class TextEnhancer:
    """Сервис для улучшения качества OCR распознавания с помощью Claude AI"""
    
    def __init__(self):
        self.claude_client = None
        if CLAUDE_ENABLED and CLAUDE_API_KEY:
            self.claude_client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    
    def analyze_text_quality(self, text: str) -> Dict[str, any]:
        """Анализирует качество распознанного текста"""
        issues = {
            'ukrainian_chars': [],
            'ocr_errors': [],
            'mixed_languages': False,
            'confidence_score': 0,
            'needs_enhancement': False
        }
        
        # Расширенный список украинских символов и их русских эквивалентов
        ukrainian_chars = [
            'ї', 'і', 'є', 'ґ', 'ў',  # основные
            'Ї', 'І', 'Є', 'Ґ', 'Ў',  # заглавные
            'ъ', 'ы', 'э',             # редко используемые в украинском
            'щ', 'ч', 'ш', 'ж',        # могут быть искажены
        ]
        
        # Проверяем наличие украинских символов
        found_ukrainian = []
        for char in ukrainian_chars:
            if char in text:
                found_ukrainian.append(char)
        
        issues['ukrainian_chars'] = found_ukrainian
        
        # Дополнительная проверка на украинские слова и конструкции
        ukrainian_patterns = [
            r'\bукраїн',  # украинский
            r'\bукраїнськ',  # украинский
            r'ський\b',   # украинские окончания
            r'цький\b',   # украинские окончания
            r'\bна\s+\d+\s+року\b',  # "на 2024 року" вместо "на 2024 год"
            r'\bрік\b',   # "рік" вместо "год"
            r'\bроку\b',  # "року" вместо "года"
        ]
        
        for pattern in ukrainian_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                issues['ukrainian_chars'].append(f"pattern:{pattern}")
        
        # Типичные ошибки OCR
        ocr_patterns = [
            (r'\b0[а-я]', 'О замененo на 0'),
            (r'[а-я]0\b', 'О замененo на 0'),
            (r'\bl[а-я]', 'I заменена на l'),
            (r'[а-я]l\b', 'I заменена на l'),
            (r'rn', 'm заменена на rn'),
            (r'\d[а-я]', 'Цифра перед буквой'),
            (r'[а-я]\d', 'Буква перед цифрой')
        ]
        
        for pattern, description in ocr_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                issues['ocr_errors'].append({
                    'pattern': pattern,
                    'description': description,
                    'count': len(matches)
                })
        
        # Проверка смешанных языков
        cyrillic_count = len(re.findall(r'[а-яё]', text, re.IGNORECASE))
        latin_count = len(re.findall(r'[a-z]', text, re.IGNORECASE))
        
        if cyrillic_count > 0 and latin_count > 0:
            ratio = min(cyrillic_count, latin_count) / max(cyrillic_count, latin_count)
            if ratio > 0.1:  # Если соотношение больше 10%
                issues['mixed_languages'] = True
        
        # Расчет оценки качества (от 0 до 100)
        quality_score = 100
        quality_score -= len(issues['ukrainian_chars']) * 15  # Увеличили штраф за украинские символы
        quality_score -= sum(err['count'] for err in issues['ocr_errors']) * 5
        if issues['mixed_languages']:
            quality_score -= 20
        
        issues['confidence_score'] = max(0, quality_score)
        # Всегда улучшаем текст, если есть украинские символы или OCR ошибки
        issues['needs_enhancement'] = len(issues['ukrainian_chars']) > 0 or len(issues['ocr_errors']) > 0
        
        return issues
    
    def preprocess_ukrainian_text(self, text: str) -> str:
        """Предварительная обработка украинского текста перед отправкой в Claude AI"""
        if not text:
            return text
            
        # Максимально расширенный словарь украинских замен
        replacements = {
            # Основные украинские символы
            'ї': 'и', 'Ї': 'И',
            'і': 'и', 'І': 'И',
            'є': 'е', 'Є': 'Е', 
            'ґ': 'г', 'Ґ': 'Г',
            'ў': 'у', 'Ў': 'У',
            
            # Дополнительные символы
            'ъ': 'ь',  # твердый знак
            'ы': 'и',  # ы не используется в украинском
            'э': 'е',  # э редко используется
            
            # Возможные искажения при OCR
            'ѐ': 'е', 'ѐ': 'е',  # е с ударением
            'ѓ': 'г', 'Ѓ': 'Г',  # г с ударением
            'ќ': 'к', 'Ќ': 'К',  # к с ударением
            
            # Основные украинские слова
            'Муніципальне': 'Муниципальное',
            'муніципальне': 'муниципальное',
            'Муніципальний': 'Муниципальный',
            'муніципальний': 'муниципальный',
            'Свідетельство': 'Свидетельство',
            'свідетельство': 'свидетельство',
            'свідоцтво': 'свидетельство',
            'Свідоцтво': 'Свидетельство',
            'ІНН': 'ИНН', 'іНН': 'ИНН', 'інн': 'ИНН',
            'КПП': 'КПП', 'кпп': 'КПП',
            'ОГРН': 'ОГРН', 'огрн': 'ОГРН',
            'БІК': 'БИК', 'бік': 'БИК',
            
            # Временные конструкции
            'року': 'года', 'Року': 'Года',
            'рік': 'год', 'Рік': 'Год',
            'місяць': 'месяц', 'Місяць': 'Месяц',
            'день': 'день', 'День': 'День',
            
            # Месяцы
            'жовтня': 'октября', 'Жовтня': 'Октября',
            'березня': 'марта', 'Березня': 'Марта',
            'квітня': 'апреля', 'Квітня': 'Апреля',
            'травня': 'мая', 'Травня': 'Мая',
            'червня': 'июня', 'Червня': 'Июня',
            'липня': 'июля', 'Липня': 'Июля',
            'серпня': 'августа', 'Серпня': 'Августа',
            'вересня': 'сентября', 'Вересня': 'Сентября',
            'листопада': 'ноября', 'Листопада': 'Ноября',
            'грудня': 'декабря', 'Грудня': 'Декабря',
            'січня': 'января', 'Січня': 'Января',
            'лютого': 'февраля', 'Лютого': 'Февраля',
            
            # Украинские окончания и суффиксы
            'реєстраційний': 'регистрационный',
            'Реєстраційний': 'Регистрационный',
            'реєстрація': 'регистрация',
            'Реєстрація': 'Регистрация',
            'українськ': 'русск', 'Українськ': 'Русск',
            'ський': 'ский', 'Ський': 'Ский',
            'цький': 'цкий', 'Цький': 'Цкий',
            'тися': 'ться', 'Тися': 'Ться',
            'ння': 'ние', 'Ння': 'Ние',
            'ення': 'ение', 'Ення': 'Ение',
            'ання': 'ание', 'Ання': 'Ание',
            'ування': 'ование', 'Ування': 'Ование',
            
            # Предлоги и союзы
            'з дня': 'с дня', 'З дня': 'С дня',
            'до дня': 'до дня', 'До дня': 'До дня',
            'від': 'от', 'Від': 'От',
            'для': 'для', 'Для': 'Для',
            'при': 'при', 'При': 'При',
            'під': 'под', 'Під': 'Под',
            'над': 'над', 'Над': 'Над',
            'через': 'через', 'Через': 'Через',
            
            # Образовательные термины
            'установа': 'учреждение', 'Установа': 'Учреждение',
            'заклад': 'учреждение', 'Заклад': 'Учреждение',
            'загальноосвітнє': 'общеобразовательное',
            'Загальноосвітнє': 'Общеобразовательное',
            'акредитації': 'аккредитации',
            'Акредитації': 'Аккредитации',
            'середня': 'средняя', 'Середня': 'Средняя',
            'гімназія': 'гимназия', 'Гімназія': 'Гимназия',
            'ліцей': 'лицей', 'Ліцей': 'Лицей',
            
            # Географические термины
            'місто': 'город', 'Місто': 'Город',
            'село': 'село', 'Село': 'Село',
            'вулиця': 'улица', 'Вулиця': 'Улица',
            'будинок': 'дом', 'Будинок': 'Дом',
            'квартира': 'квартира', 'Квартира': 'Квартира',
            
            # Банковские термины
            'рахунок': 'счет', 'Рахунок': 'Счет',
            'розрахунковий': 'расчетный', 'Розрахунковий': 'Расчетный',
            'кореспондентський': 'корреспондентский',
            'Кореспондентський': 'Корреспондентский',
            
            # Конкретные названия
            'Волгоградська': 'Волгоградская',
            'Серафимовичський': 'Серафимовичский',
            'Бобровська': 'Бобровская',
            'Центральна': 'Центральная',
            'казенне': 'казенное', 'Казенне': 'Казенное',
            'державне': 'государственное', 'Державне': 'Государственное',
            
            # Должности
            'керівник': 'руководитель', 'Керівник': 'Руководитель',
            'завідувач': 'заведующий', 'Завідувач': 'Заведующий',
            'працівник': 'работник', 'Працівник': 'Работник',
            'співробітник': 'сотрудник', 'Співробітник': 'Сотрудник',
        }
        
        processed_text = text
        
        # Применяем все замены
        for ukr, rus in replacements.items():
            processed_text = processed_text.replace(ukr, rus)
        
        # Дополнительные паттерны с регулярными выражениями
        import re
        patterns = [
            # Временные конструкции
            (r'\bна\s+(\d+)\s+року\b', r'на \1 года'),
            (r'\bу\s+(\d+)\s+році\b', r'в \1 году'),
            (r'\b(\d+)\s+року\b', r'\1 года'),
            (r'\b(\d+)\s+рік\b', r'\1 год'),
            
            # Украинские предлоги с числами
            (r'\bз\s+(\d+)', r'с \1'),
            (r'\bдо\s+(\d+)', r'до \1'),
            (r'\bвід\s+(\d+)', r'от \1'),
            
            # Украинские падежные окончания
            (r'([а-яё]+)ський\b', r'\1ский'),
            (r'([а-яё]+)цький\b', r'\1цкий'),
            (r'([а-яё]+)ння\b', r'\1ние'),
            (r'([а-яё]+)ення\b', r'\1ение'),
            (r'([а-яё]+)ання\b', r'\1ание'),
            (r'([а-яё]+)ування\b', r'\1ование'),
        ]
        
        for pattern, replacement in patterns:
            processed_text = re.sub(pattern, replacement, processed_text, flags=re.IGNORECASE)
        
        return processed_text

    async def enhance_russian_text(self, text: str, context: str = "") -> str:
        """Улучшает русский текст с помощью Claude AI"""
        if not self.claude_client:
            logger.warning("Claude AI not available for text enhancement")
            return text
        
        # Предварительная обработка украинского текста
        preprocessed_text = self.preprocess_ukrainian_text(text)
        
        # Создаем промпт для Claude AI с максимально агрессивными инструкциями
        prompt = f"""
Ты - эксперт по исправлению OCR-ошибок и переводу украинского текста на русский язык.

КРИТИЧЕСКИ ВАЖНО: Этот документ ДОЛЖЕН быть полностью на русском языке!

ТВОЯ ЗАДАЧА:
1. ОБЯЗАТЕЛЬНО переведи ВСЕ украинские слова на русский язык
2. Исправь ВСЕ OCR-ошибки и искажения текста
3. Восстанови правильную структуру и форматирование
4. Сохрани все числа, даты и коды точно как есть

УКРАИНСКИЕ СИМВОЛЫ → РУССКИЕ (ОБЯЗАТЕЛЬНО ЗАМЕНИТЬ):
- ї, і → и
- є → е  
- ґ → г
- ў → у
- Все украинские буквы должны стать русскими!

УКРАИНСКИЕ СЛОВА → РУССКИЕ (ПРИМЕРЫ):
- Муніципальне → Муниципальное
- Свідетельство → Свидетельство
- ІНН → ИНН
- року → года
- рік → год
- реєстраційний → регистрационный
- установа/заклад → учреждение
- загальноосвітнє → общеобразовательное
- акредитації → аккредитации
- середня → средняя
- гімназія → гимназия
- ліцей → лицей
- місто → город
- вулиця → улица
- будинок → дом
- рахунок → счет
- розрахунковий → расчетный
- директор/керівник → директор/руководитель

УКРАИНСКИЕ КОНСТРУКЦИИ:
- "на XXXX року" → "на XXXX года"
- "у XXXX році" → "в XXXX году"  
- "з дня" → "с дня"
- "від" → "от"
- окончания "-ський" → "-ский"
- окончания "-цький" → "-цкий"
- окончания "-ння" → "-ние"

ИСПРАВЬ ТИПИЧНЫЕ OCR-ОШИБКИ:
- Замени похожие символы (0→О, 1→I, rn→m, и т.д.)
- Восстанови пропущенные буквы
- Исправь разорванные слова
- Убери лишние пробелы и символы

ВАЖНО: Результат должен быть ТОЛЬКО на русском языке, без украинских слов!

Исходные данные для обработки:
{preprocessed_text}

Верни ТОЛЬКО исправленный текст в том же формате (строка за строкой), без дополнительных комментариев."""

        try:
            response = await asyncio.to_thread(
                self.claude_client.messages.create,
                model=CLAUDE_MODEL,
                max_tokens=8192,
                messages=[{"role": "user", "content": prompt}]
            )
            
            enhanced_text = response.content[0].text.strip()
            logger.info("Text enhanced successfully with Claude AI")
            return enhanced_text
            
        except Exception as e:
            logger.error(f"Claude AI enhancement failed: {e}")
            return text
    
    async def process_xlsx_file(self, xlsx_data: bytes, file_name: str = "") -> Optional[bytes]:
        """Обрабатывает XLSX файл, улучшая качество текста"""
        if not CLAUDE_ENABLED:
            logger.info("Claude AI not enabled, returning original file")
            return xlsx_data
        
        try:
            # Загружаем файл
            workbook = openpyxl.load_workbook(BytesIO(xlsx_data))
            enhancement_performed = False
            
            logger.info(f"Processing {len(workbook.sheetnames)} sheets for text enhancement")
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"Analyzing sheet: {sheet_name}")
                
                # Собираем все ячейки с текстом
                text_cells = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value and isinstance(cell.value, str):
                            text_cells.append({
                                'row': row_idx,
                                'col': col_idx,
                                'value': cell.value,
                                'cell': cell
                            })
                
                if not text_cells:
                    logger.info(f"No text cells found in sheet {sheet_name}")
                    continue
                
                # Анализируем качество текста для листа
                all_text = " ".join([cell['value'] for cell in text_cells])
                quality_analysis = self.analyze_text_quality(all_text)
                
                logger.info(f"Sheet {sheet_name} quality score: {quality_analysis['confidence_score']}")
                
                # Обрабатываем ВСЕ листы независимо от качества
                logger.info(f"Enhancing sheet {sheet_name} with {len(text_cells)} text cells")
                
                # Группируем ячейки для обработки (по 30 ячеек для соблюдения лимита токенов)
                batch_size = 30
                for i in range(0, len(text_cells), batch_size):
                    batch = text_cells[i:i + batch_size]
                    
                    # Объединяем текст из ячеек
                    batch_text = "\n".join([f"Ячейка {cell['row']},{cell['col']}: {cell['value']}" 
                                          for cell in batch])
                    
                    # Улучшаем текст
                    enhanced_text = await self.enhance_russian_text(
                        batch_text, 
                        f"Таблица '{sheet_name}' из файла '{file_name}'"
                    )
                    
                    # Разбираем результат и обновляем ячейки
                    enhanced_lines = enhanced_text.split('\n')
                    for j, line in enumerate(enhanced_lines):
                        if j >= len(batch):
                            break
                        
                        # Извлекаем новое значение ячейки
                        if ': ' in line:
                            new_value = line.split(': ', 1)[1].strip()
                            if new_value and new_value != batch[j]['value']:
                                batch[j]['cell'].value = new_value
                                enhancement_performed = True
                                logger.debug(f"Enhanced cell {batch[j]['row']},{batch[j]['col']}: "
                                           f"'{batch[j]['value']}' -> '{new_value}'")
                
                logger.info(f"Completed enhancement for sheet {sheet_name}")
            
            if enhancement_performed:
                # Сохраняем улучшенный файл
                output_buffer = BytesIO()
                workbook.save(output_buffer)
                enhanced_data = output_buffer.getvalue()
                logger.info("XLSX file enhancement completed successfully")
                return enhanced_data
            else:
                logger.info("No enhancement was needed")
                return xlsx_data
                
        except ImportError as e:
            logger.error(f"Required package not available: {e}")
            return xlsx_data
        except Exception as e:
            logger.error(f"Error processing XLSX file: {e}")
            return xlsx_data
    
    def get_enhancement_stats(self, original_text: str, enhanced_text: str) -> Dict[str, any]:
        """Возвращает статистику улучшений"""
        original_analysis = self.analyze_text_quality(original_text)
        enhanced_analysis = self.analyze_text_quality(enhanced_text)
        
        return {
            'original_score': original_analysis['confidence_score'],
            'enhanced_score': enhanced_analysis['confidence_score'],
            'improvement': enhanced_analysis['confidence_score'] - original_analysis['confidence_score'],
            'ukrainian_chars_fixed': len(original_analysis['ukrainian_chars']),
            'ocr_errors_fixed': len(original_analysis['ocr_errors'])
        } 