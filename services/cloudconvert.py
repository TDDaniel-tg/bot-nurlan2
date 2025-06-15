import aiohttp
import asyncio
import logging
from typing import Optional, Dict, Any, BinaryIO
from config.settings import (
    CLOUDCONVERT_API_KEY, 
    CLOUDCONVERT_BASE_URL, 
    CLOUDCONVERT_OCR_LANGUAGES,
    CLOUDCONVERT_LOCALE,
    API_TIMEOUT,
    CLAUDE_ENABLED,
    CLAUDE_API_KEY,
    CLAUDE_MODEL
)

logger = logging.getLogger(__name__)

class CloudConvertService:
    def __init__(self):
        self.api_key = CLOUDCONVERT_API_KEY
        self.base_url = CLOUDCONVERT_BASE_URL
        self.headers = {
            'Authorization': f'Bearer {self.api_key}',
            'Content-Type': 'application/json'
        }
    
    async def create_conversion_job(self, file_name: str) -> Optional[Dict[str, Any]]:
        """Создание задачи конвертации с принудительными русскими настройками"""
        job_payload = {
            "tasks": {
                "import-pdf": {
                    "operation": "import/upload"
                },
                "convert-to-xlsx": {
                    "operation": "convert",
                    "input": "import-pdf",
                    "input_format": "pdf",
                    "output_format": "xlsx",
                    "options": {
                        "ocr_lang": ["rus"],                      # ТОЛЬКО русский язык
                        "ocr_accuracy": "best",                   # Лучшее качество OCR
                        "locale": "ru_RU",                       # Принудительно русская локаль
                        "text_encoding": "utf-8"                 # UTF-8 кодировка
                    }
                },
                "export-xlsx": {
                    "operation": "export/url",
                    "input": "convert-to-xlsx"
                }
            }
        }
        
        try:
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=API_TIMEOUT)) as session:
                async with session.post(
                    f"{self.base_url}/jobs",
                    json=job_payload,
                    headers=self.headers
                ) as response:
                    if response.status == 201:
                        job_data = await response.json()
                        logger.info(f"Created conversion job {job_data['data']['id']} for file {file_name}")
                        logger.debug(f"Job data structure: {job_data['data']}")
                        return job_data['data']
                    else:
                        error_text = await response.text()
                        if response.status == 402:
                            logger.error(f"CloudConvert credits exceeded: {error_text}")
                        else:
                            logger.error(f"Failed to create job: {response.status} - {error_text}")
                        return None
        
        except asyncio.TimeoutError:
            logger.error("Timeout while creating conversion job")
            return None
        except Exception as e:
            logger.error(f"Error creating conversion job: {e}")
            return None
    
    async def upload_file(self, upload_data: dict, file_data: bytes, file_name: str) -> bool:
        """Загрузка файла в CloudConvert используя параметры формы"""
        try:
            upload_url = upload_data.get('url')
            upload_parameters = upload_data.get('parameters', {})
            
            if not upload_url:
                logger.error("No upload URL provided")
                return False
            
            # Создаем форму с параметрами от CloudConvert
            form_data = aiohttp.FormData()
            
            # Добавляем все параметры, которые требует CloudConvert
            for key, value in upload_parameters.items():
                form_data.add_field(key, value)
            
            # Добавляем файл (обычно это поле называется 'file')
            form_data.add_field('file', file_data, filename=file_name, content_type='application/pdf')
            
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=API_TIMEOUT)) as session:
                async with session.post(upload_url, data=form_data) as response:
                    if response.status in [200, 201, 204]:
                        logger.info(f"Successfully uploaded file {file_name}")
                        return True
                    else:
                        error_text = await response.text()
                        logger.error(f"Failed to upload file: {response.status} - {error_text}")
                        return False
        
        except asyncio.TimeoutError:
            logger.error("Timeout while uploading file")
            return False
        except Exception as e:
            logger.error(f"Error uploading file: {e}")
            return False
    
    async def get_job_status(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Получение статуса задачи"""
        try:
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=API_TIMEOUT)) as session:
                async with session.get(
                    f"{self.base_url}/jobs/{job_id}",
                    headers=self.headers
                ) as response:
                    if response.status == 200:
                        job_data = await response.json()
                        return job_data['data']
                    else:
                        error_text = await response.text()
                        logger.error(f"Failed to get job status: {response.status} - {error_text}")
                        return None
        
        except asyncio.TimeoutError:
            logger.error("Timeout while getting job status")
            return None
        except Exception as e:
            logger.error(f"Error getting job status: {e}")
            return None
    
    async def download_file(self, download_url: str) -> Optional[bytes]:
        """Скачивание конвертированного файла"""
        try:
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=API_TIMEOUT)) as session:
                async with session.get(download_url) as response:
                    if response.status == 200:
                        file_data = await response.read()
                        logger.info("Successfully downloaded converted file")
                        return file_data
                    else:
                        error_text = await response.text()
                        logger.error(f"Failed to download file: {response.status} - {error_text}")
                        return None
        
        except asyncio.TimeoutError:
            logger.error("Timeout while downloading file")
            return None
        except Exception as e:
            logger.error(f"Error downloading file: {e}")
            return None
    
    async def wait_for_completion(self, job_id: str, max_wait_time: int = 300) -> Optional[str]:
        """Ожидание завершения конвертации с возвращением URL для скачивания"""
        start_time = asyncio.get_event_loop().time()
        
        while True:
            # Проверяем таймаут
            if asyncio.get_event_loop().time() - start_time > max_wait_time:
                logger.error(f"Conversion timeout for job {job_id}")
                return None
            
            job_status = await self.get_job_status(job_id)
            if not job_status:
                return None
            
            status = job_status.get('status')
            logger.info(f"Job {job_id} status: {status}")
            
            if status == 'finished':
                # Ищем задачу экспорта для получения URL
                tasks = job_status.get('tasks', [])
                if isinstance(tasks, dict):
                    # Если tasks - словарь
                    for task_id, task in tasks.items():
                        if task.get('operation') == 'export/url' and task.get('status') == 'finished':
                            download_url = task.get('result', {}).get('files', [{}])[0].get('url')
                            if download_url:
                                return download_url
                else:
                    # Если tasks - список
                    for task in tasks:
                        if task.get('operation') == 'export/url' and task.get('status') == 'finished':
                            download_url = task.get('result', {}).get('files', [{}])[0].get('url')
                            if download_url:
                                return download_url
                
                logger.error(f"Job {job_id} finished but no download URL found")
                return None
            
            elif status == 'error':
                # Получаем детальную информацию об ошибке
                error_message = job_status.get('message', 'Unknown error')
                logger.error(f"Job {job_id} failed with error: {error_message}")
                
                # Проверяем ошибки в отдельных задачах
                tasks = job_status.get('tasks', [])
                if isinstance(tasks, dict):
                    for task_id, task in tasks.items():
                        if task.get('status') == 'error':
                            task_error = task.get('message', 'Unknown task error')
                            logger.error(f"Task {task_id} ({task.get('operation')}) failed: {task_error}")
                else:
                    for task in tasks:
                        if task.get('status') == 'error':
                            task_error = task.get('message', 'Unknown task error')
                            logger.error(f"Task {task.get('name', 'unknown')} ({task.get('operation')}) failed: {task_error}")
                
                return None
            
            # Ждем перед следующей проверкой
            await asyncio.sleep(5)
    
    async def force_ukrainian_to_russian_conversion(self, xlsx_data: bytes, file_name: str) -> bytes:
        """Принудительная замена украинских символов на русские в XLSX файле"""
        try:
            import openpyxl
            from io import BytesIO
            import re
            
            # Максимально расширенный словарь замен украинских символов
            ukrainian_to_russian = {
                # Основные украинские символы
                'ї': 'и', 'Ї': 'И',
                'і': 'и', 'І': 'И', 
                'є': 'е', 'Є': 'Е',
                'ґ': 'г', 'Ґ': 'Г',
                'ў': 'у', 'Ў': 'У',
                
                # Дополнительные символы
                'ъ': 'ь',  # твердый знак
                'ы': 'и',  # ы не используется в украинском
                'э': 'е',  # э редко используется
                
                # Возможные искажения при OCR
                'ѐ': 'е', 'ѐ': 'е',  # е с ударением
                'ѓ': 'г', 'Ѓ': 'Г',  # г с ударением
                'ќ': 'к', 'Ќ': 'К',  # к с ударением
            }
            
            # Максимально расширенный словарь украинских слов и конструкций
            ukrainian_words = {
                # Основные административные термины
                'Муніципальне': 'Муниципальное',
                'муніципальне': 'муниципальное',
                'Муніципальний': 'Муниципальный',
                'муніципальний': 'муниципальный',
                'Свідетельство': 'Свидетельство',
                'свідетельство': 'свидетельство',
                'свідоцтво': 'свидетельство',
                'Свідоцтво': 'Свидетельство',
                
                # Документооборот
                'ІНН': 'ИНН', 'іНН': 'ИНН', 'інн': 'ИНН',
                'КПП': 'КПП', 'кпп': 'КПП',
                'ОГРН': 'ОГРН', 'огрн': 'ОГРН',
                'БІК': 'БИК', 'бік': 'БИК',
                'реєстраційний': 'регистрационный',
                'Реєстраційний': 'Регистрационный',
                'реєстрація': 'регистрация',
                'Реєстрація': 'Регистрация',
                
                # Временные конструкции
                'року': 'года', 'Року': 'Года',
                'рік': 'год', 'Рік': 'Год',
                'місяць': 'месяц', 'Місяць': 'Месяц',
                'день': 'день', 'День': 'День',
                'жовтня': 'октября', 'Жовтня': 'Октября',
                'березня': 'марта', 'Березня': 'Марта',
                'квітня': 'апреля', 'Квітня': 'Апреля',
                'травня': 'мая', 'Травня': 'Мая',
                'червня': 'июня', 'Червня': 'Июня',
                'липня': 'июля', 'Липня': 'Июля',
                'серпня': 'августа', 'Серпня': 'Августа',
                'вересня': 'сентября', 'Вересня': 'Сентября',
                'листопада': 'ноября', 'Листопада': 'Ноября',
                'грудня': 'декабря', 'Грудня': 'Декабря',
                'січня': 'января', 'Січня': 'Января',
                'лютого': 'февраля', 'Лютого': 'Февраля',
                
                # Украинские окончания и суффиксы
                'українськ': 'русск', 'Українськ': 'Русск',
                'ський': 'ский', 'Ський': 'Ский',
                'цький': 'цкий', 'Цький': 'Цкий',
                'тися': 'ться', 'Тися': 'Ться',
                'ння': 'ние', 'Ння': 'Ние',
                'ення': 'ение', 'Ення': 'Ение',
                'ання': 'ание', 'Ання': 'Ание',
                'ування': 'ование', 'Ування': 'Ование',
                
                # Предлоги и союзы
                'з дня': 'с дня', 'З дня': 'С дня',
                'до дня': 'до дня', 'До дня': 'До дня',
                'від': 'от', 'Від': 'От',
                'для': 'для', 'Для': 'Для',
                'при': 'при', 'При': 'При',
                'під': 'под', 'Під': 'Под',
                'над': 'над', 'Над': 'Над',
                'через': 'через', 'Через': 'Через',
                
                # Образовательные термины
                'учреждение': 'учреждение',
                'установа': 'учреждение', 'Установа': 'Учреждение',
                'заклад': 'учреждение', 'Заклад': 'Учреждение',
                'общеобразовательное': 'общеобразовательное',
                'загальноосвітнє': 'общеобразовательное',
                'Загальноосвітнє': 'Общеобразовательное',
                'аккредитации': 'аккредитации',
                'акредитації': 'аккредитации',
                'Акредитації': 'Аккредитации',
                'школа': 'школа', 'Школа': 'Школа',
                'середня': 'средняя', 'Середня': 'Средняя',
                'гімназія': 'гимназия', 'Гімназія': 'Гимназия',
                'ліцей': 'лицей', 'Ліцей': 'Лицей',
                
                # Географические термины
                'область': 'область', 'Область': 'Область',
                'район': 'район', 'Район': 'Район',
                'місто': 'город', 'Місто': 'Город',
                'село': 'село', 'Село': 'Село',
                'вулиця': 'улица', 'Вулиця': 'Улица',
                'будинок': 'дом', 'Будинок': 'Дом',
                'квартира': 'квартира', 'Квартира': 'Квартира',
                
                # Банковские термины
                'банк': 'банк', 'Банк': 'Банк',
                'рахунок': 'счет', 'Рахунок': 'Счет',
                'розрахунковий': 'расчетный', 'Розрахунковий': 'Расчетный',
                'кореспондентський': 'корреспондентский',
                'Кореспондентський': 'Корреспондентский',
                
                # Конкретные названия из документа
                'Волгоградська': 'Волгоградская',
                'Серафимовичський': 'Серафимовичский',
                'Бобровська': 'Бобровская',
                'Центральна': 'Центральная',
                'казенне': 'казенное', 'Казенне': 'Казенное',
                'державне': 'государственное', 'Державне': 'Государственное',
                
                # Дополнительные часто встречающиеся слова
                'директор': 'директор', 'Директор': 'Директор',
                'керівник': 'руководитель', 'Керівник': 'Руководитель',
                'завідувач': 'заведующий', 'Завідувач': 'Заведующий',
                'працівник': 'работник', 'Працівник': 'Работник',
                'співробітник': 'сотрудник', 'Співробітник': 'Сотрудник',
            }
            
            # Загружаем XLSX файл
            workbook = openpyxl.load_workbook(BytesIO(xlsx_data))
            replacements_made = 0
            
            logger.info(f"Принудительная замена украинских символов в файле {file_name}")
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            original_value = cell.value
                            new_value = original_value
                            
                            # Заменяем украинские символы
                            for ukr_char, rus_char in ukrainian_to_russian.items():
                                new_value = new_value.replace(ukr_char, rus_char)
                            
                            # Заменяем украинские слова (точная замена)
                            for ukr_word, rus_word in ukrainian_words.items():
                                new_value = new_value.replace(ukr_word, rus_word)
                            
                            # Заменяем украинские слова с границами слов (регулярные выражения)
                            for ukr_word, rus_word in ukrainian_words.items():
                                pattern = r'\b' + re.escape(ukr_word) + r'\b'
                                new_value = re.sub(pattern, rus_word, new_value, flags=re.IGNORECASE)
                            
                            # Дополнительные паттерны для украинских конструкций
                            patterns = [
                                # Временные конструкции
                                (r'\bна\s+(\d+)\s+року\b', r'на \1 года'),
                                (r'\bу\s+(\d+)\s+році\b', r'в \1 году'),
                                (r'\b(\d+)\s+року\b', r'\1 года'),
                                (r'\b(\d+)\s+рік\b', r'\1 год'),
                                
                                # Украинские предлоги
                                (r'\bз\s+(\d+)', r'с \1'),
                                (r'\bдо\s+(\d+)', r'до \1'),
                                (r'\bвід\s+(\d+)', r'от \1'),
                                
                                # Украинские падежные окончания
                                (r'([а-яё]+)ський\b', r'\1ский'),
                                (r'([а-яё]+)цький\b', r'\1цкий'),
                                (r'([а-яё]+)ння\b', r'\1ние'),
                                (r'([а-яё]+)ення\b', r'\1ение'),
                                (r'([а-яё]+)ання\b', r'\1ание'),
                            ]
                            
                            for pattern, replacement in patterns:
                                new_value = re.sub(pattern, replacement, new_value, flags=re.IGNORECASE)
                            
                            # Если были изменения, обновляем ячейку
                            if new_value != original_value:
                                cell.value = new_value
                                replacements_made += 1
                                logger.debug(f"Заменено: '{original_value}' → '{new_value}'")
            
            if replacements_made > 0:
                logger.info(f"Выполнено {replacements_made} замен украинских символов/слов")
                
                # Сохраняем обновленный файл
                output_buffer = BytesIO()
                workbook.save(output_buffer)
                return output_buffer.getvalue()
            else:
                logger.info("Украинские символы не найдены")
                return xlsx_data
                
        except Exception as e:
            logger.error(f"Ошибка при принудительной замене украинских символов: {e}")
            return xlsx_data

    async def enhance_text_with_claude(self, xlsx_data: bytes, file_name: str = "") -> Optional[bytes]:
        """Улучшение качества распознанного текста с помощью TextEnhancer"""
        try:
            # Сначала принудительно заменяем украинские символы
            xlsx_data = await self.force_ukrainian_to_russian_conversion(xlsx_data, file_name)
            
            # Затем применяем Claude AI если доступен
            if CLAUDE_ENABLED:
                from services.text_enhancer import TextEnhancer
                enhancer = TextEnhancer()
                enhanced_data = await enhancer.process_xlsx_file(xlsx_data, file_name)
                return enhanced_data
            else:
                logger.info("Claude AI не включен, используется только принудительная замена")
                return xlsx_data
            
        except ImportError:
            logger.warning("TextEnhancer not available, using only forced replacement")
            return xlsx_data
        except Exception as e:
            logger.error(f"Error in text enhancement: {e}")
            return xlsx_data

    async def convert_pdf_to_xlsx(self, file_data: bytes, file_name: str) -> Optional[bytes]:
        """Полный процесс конвертации PDF в XLSX с улучшением качества и множественными попытками"""
        try:
            logger.info(f"Starting conversion of {file_name} with improved quality settings")
            
            # 1. Первая попытка: Улучшенная двухэтапная конвертация PDF→DOCX→XLSX
            job_data = await self.create_conversion_job(file_name)
            if job_data:
                result = await self._process_conversion_job(job_data, file_data, file_name, "standard enhanced")
                if result:
                    return result
            
            # 2. Вторая попытка: Альтернативная стратегия PDF→CSV→XLSX
            logger.info(f"Trying alternative high-quality conversion for {file_name}")
            job_data = await self.create_high_quality_conversion_job(file_name)
            if job_data:
                result = await self._process_conversion_job(job_data, file_data, file_name, "high-quality CSV")
                if result:
                    return result
            
            logger.error(f"All conversion attempts failed for {file_name}")
            return None
            
        except Exception as e:
            logger.error(f"Error in conversion process: {e}")
            return None
    
    async def _process_conversion_job(self, job_data: Dict[str, Any], file_data: bytes, file_name: str, strategy: str) -> Optional[bytes]:
        """Обработка задачи конвертации с указанной стратегией"""
        try:
            job_id = job_data['id']
            logger.info(f"Processing conversion job {job_id} with {strategy} strategy")
            
            # 2. Получаем данные для загрузки файла из задачи импорта
            upload_data = None
            
            tasks = job_data.get('tasks', [])
            if isinstance(tasks, dict):
                # Если tasks - словарь, итерируем по items()
                for task_id, task in tasks.items():
                    if task.get('operation') == 'import/upload':
                        upload_data = task.get('result', {}).get('form', {})
                        break
            else:
                # Если tasks - список, итерируем напрямую
                for task in tasks:
                    if task.get('operation') == 'import/upload':
                        upload_data = task.get('result', {}).get('form', {})
                        break
            
            if not upload_data or not upload_data.get('url'):
                logger.error("No upload data found in job tasks")
                # Попробуем получить данные напрямую из API
                job_status = await self.get_job_status(job_id)
                if job_status:
                    tasks = job_status.get('tasks', [])
                    if isinstance(tasks, dict):
                        for task_id, task in tasks.items():
                            if task.get('operation') == 'import/upload' and task.get('status') == 'waiting':
                                upload_data = task.get('result', {}).get('form', {})
                                break
                    else:
                        for task in tasks:
                            if task.get('operation') == 'import/upload' and task.get('status') == 'waiting':
                                upload_data = task.get('result', {}).get('form', {})
                                break
                
                if not upload_data or not upload_data.get('url'):
                    logger.error(f"Could not find upload data for {strategy} strategy")
                    return None
            
            # 3. Загружаем файл
            upload_success = await self.upload_file(upload_data, file_data, file_name)
            if not upload_success:
                logger.error(f"Upload failed for {strategy} strategy")
                return None
            
            # 4. Ждем завершения конвертации
            download_url = await self.wait_for_completion(job_id)
            if not download_url:
                logger.error(f"Conversion failed for {strategy} strategy")
                return None
            
            # 5. Скачиваем результат
            converted_file = await self.download_file(download_url)
            if not converted_file:
                logger.error(f"Download failed for {strategy} strategy")
                return None
            
            logger.info(f"Successfully converted {file_name} using {strategy} strategy")
            
            # 6. Улучшаем качество с помощью Claude AI (если включен)
            enhanced_file = await self.enhance_text_with_claude(converted_file, file_name)
            logger.info(f"Text enhancement completed for {file_name}")
            return enhanced_file
            
        except Exception as e:
            logger.error(f"Error in {strategy} conversion process: {e}")
            return None

    async def create_high_quality_conversion_job(self, file_name: str) -> Optional[Dict[str, Any]]:
        """Создание задачи конвертации с альтернативными OCR настройками"""
        job_payload = {
            "tasks": {
                "import-pdf": {
                    "operation": "import/upload"
                },
                "convert-to-xlsx": {
                    "operation": "convert",
                    "input": "import-pdf",
                    "input_format": "pdf",
                    "output_format": "xlsx",
                    "options": {
                        "ocr_lang": ["rus"],                      # Только русский язык
                        "ocr_accuracy": "fast",                   # Быстрое распознавание
                        "ocr_engine": "tesseract"                 # Tesseract OCR
                    }
                },
                "export-xlsx": {
                    "operation": "export/url",
                    "input": "convert-to-xlsx"
                }
            }
        }
        
        try:
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=API_TIMEOUT)) as session:
                async with session.post(
                    f"{self.base_url}/jobs",
                    json=job_payload,
                    headers=self.headers
                ) as response:
                    if response.status == 201:
                        job_data = await response.json()
                        logger.info(f"Created high-quality conversion job {job_data['data']['id']} for file {file_name}")
                        return job_data['data']
                    else:
                        error_text = await response.text()
                        logger.error(f"Failed to create high-quality job: {response.status} - {error_text}")
                        return None
        
        except Exception as e:
            logger.error(f"Error creating high-quality conversion job: {e}")
            return None 